#!/usr/bin/env python3
"""
测试用例Excel格式转换器
将Markdown格式的测试用例文档转换为Excel表格格式

支持两种文档结构：

结构A（模块 → 用例）：
    #### TC_MOD_001 - 用例标题
    ...

结构B（模块 → 功能组 → 用例）：
    #### 功能组名称（如字段名）
    > 元数据行（引用块，会被忽略）
    ##### TC_MOD_001 - 用例标题
    ...
    结构B下所属模块列展示为：「模块名 > 功能组名」

每条用例格式：
    1. **用例信息**
       1. 优先级：P0
       2. 粒度：单功能
       3. 类型：正向
       4. 关联用户故事：...

    2. **Setup**
       - 前置条件1
       - 前置条件2

    3. **执行步骤**
       1. 步骤一
          - 预期结果1
       2. 步骤二
          - 预期结果2

    4. **Teardown**
       - 后置条件
"""
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def format_steps(step_lines):
    """
    将收集到的执行步骤行格式化为文本：
      1. 步骤描述
         → 预期结果1
         → 预期结果2
      2. 步骤描述（无预期结果）
    """
    result = []
    step_num = 0
    current_step = None
    current_expects = []

    for line in step_lines:
        stripped = line.strip()
        if not stripped:
            continue

        # 步骤行：stripped 以数字+句点开头（如 "1. 步骤内容"）
        step_m = re.match(r'^(\d+)\.\s+(.+)', stripped)
        # 预期结果行：stripped 以 "- " 开头
        is_expect = stripped.startswith('- ')

        if step_m and not is_expect:
            # 保存上一步骤
            if current_step is not None:
                result.append(current_step)
                for exp in current_expects:
                    result.append(f'   → {exp}')
            step_num += 1
            current_step = f'{step_num}. {step_m.group(2)}'
            current_expects = []
        elif is_expect and current_step is not None:
            current_expects.append(stripped[2:].strip())

    # 保存最后一个步骤
    if current_step is not None:
        result.append(current_step)
        for exp in current_expects:
            result.append(f'   → {exp}')

    return '\n'.join(result)


def parse_test_case_markdown(markdown_file):
    """解析新格式的测试用例Markdown文档"""
    with open(markdown_file, 'r', encoding='utf-8') as f:
        content = f.read()

    test_cases = []
    lines = content.split('\n')

    current_module = ''
    current_group  = ''   # 结构B：#### 级功能组名（非 TC_ 标题）
    current_tc = None
    current_section = None   # 'info' | 'setup' | 'steps' | 'teardown'
    setup_items = []
    step_collector = []
    teardown_items = []

    def flush_steps():
        """将收集的步骤行写入当前用例"""
        if current_tc is not None and step_collector:
            current_tc['执行步骤'] = format_steps(step_collector)

    def save_current_tc():
        nonlocal current_tc, step_collector
        flush_steps()
        if current_tc is not None:
            test_cases.append(current_tc)
        current_tc = None
        step_collector = []

    def make_tc(tc_id, tc_title):
        """创建用例字典，所属模块根据是否有功能组决定格式"""
        module_label = f'{current_module} > {current_group}' if current_group else current_module
        return {
            '用例ID':   tc_id,
            '用例标题': tc_title,
            '所属模块': module_label,
            '优先级':   '',
            '粒度':     '',
            '类型':     '',
            '前置条件': '',
            '执行步骤': '',
            '后置条件': '',
        }

    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        # ── 模块标题：### 模块X：名称（恰好3个#）─────────────────────
        if re.match(r'^#{3}\s', stripped) and not re.match(r'^#{4,}', stripped):
            current_module = stripped.lstrip('#').strip()
            current_group  = ''   # 新模块时重置功能组

        # ── 恰好4个#：可能是结构A的TC，也可能是结构B的功能组标题 ────
        elif re.match(r'^#{4}\s', stripped) and not re.match(r'^#{5,}', stripped):
            if re.match(r'^#{4}\s+TC_', stripped):
                # 结构A：#### TC_XXX_001 - 标题
                save_current_tc()
                current_group = ''
                match = re.match(r'^#{4}\s+(TC_\S+)\s+-\s+(.+)', stripped)
                if match:
                    current_tc = make_tc(match.group(1), match.group(2))
                    current_section = None
                    setup_items = []
                    step_collector = []
                    teardown_items = []
            else:
                # 结构B：#### 功能组名称（非TC标题，保存前一个用例并记录组名）
                save_current_tc()
                current_group = stripped.lstrip('#').strip()

        # ── 恰好5个#：结构B的TC标题 ##### TC_XXX_001 - 标题 ──────────
        elif re.match(r'^#{5}\s+TC_', stripped) and not re.match(r'^#{6,}', stripped):
            save_current_tc()
            match = re.match(r'^#{5}\s+(TC_\S+)\s+-\s+(.+)', stripped)
            if match:
                current_tc = make_tc(match.group(1), match.group(2))
                current_section = None
                setup_items = []
                step_collector = []
                teardown_items = []

        # ── TC 内容 ──────────────────────────────────────────────────
        elif current_tc is not None:

            # 小节标题检测（通过加粗关键字）
            if '**用例信息**' in stripped:
                current_section = 'info'

            elif '**Setup**' in stripped:
                current_section = 'setup'
                setup_items = []

            elif '**执行步骤**' in stripped:
                current_section = 'steps'
                step_collector = []

            elif '**Teardown**' in stripped:
                flush_steps()
                step_collector = []
                current_section = 'teardown'
                teardown_items = []

            # ── 小节内容 ──────────────────────────────────────────────
            elif current_section == 'info' and stripped:
                # 匹配形如 "1. 优先级：P0" 的子项（去掉缩进后）
                m = re.match(r'^\d+\.\s+(.+)', stripped)
                if m:
                    item = m.group(1)
                    if re.match(r'优先级[：:]', item):
                        current_tc['优先级'] = re.split(r'[：:]', item, 1)[1].strip()
                    elif re.match(r'粒度[：:]', item):
                        current_tc['粒度'] = re.split(r'[：:]', item, 1)[1].strip()
                    elif re.match(r'类型[：:]', item):
                        current_tc['类型'] = re.split(r'[：:]', item, 1)[1].strip()

            elif current_section == 'setup' and stripped.startswith('- '):
                setup_items.append(stripped[2:].strip())
                current_tc['前置条件'] = '\n'.join(setup_items)

            elif current_section == 'steps' and stripped and stripped != '---':
                step_collector.append(line)

            elif current_section == 'teardown' and stripped.startswith('- '):
                item = stripped[2:].strip()
                if item and item != '无':
                    teardown_items.append(item)
                current_tc['后置条件'] = '\n'.join(teardown_items) if teardown_items else '无'

        i += 1

    # 保存最后一个用例
    if current_tc is not None:
        flush_steps()
        test_cases.append(current_tc)

    return test_cases


def create_test_case_excel(test_cases, output_file):
    """创建测试用例Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = '测试用例'

    # ── 样式定义 ─────────────────────────────────────────────────────
    header_fill   = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font   = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)

    module_font   = Font(name='微软雅黑', size=10, bold=True, color='4472C4')
    default_font  = Font(name='微软雅黑', size=10)
    center_align  = Alignment(horizontal='center', vertical='top', wrap_text=True)
    default_align = Alignment(horizontal='left',   vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )

    fill_p0 = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    fill_p1 = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    fill_p2 = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    # ── 列定义 ───────────────────────────────────────────────────────
    # (列字母, 表头, 宽度, key)
    columns = [
        ('A', '用例ID',           18, '用例ID'),
        ('B', '用例标题',         42, '用例标题'),
        ('C', '优先级',           10, '优先级'),
        ('D', '粒度',             12, '粒度'),
        ('E', '类型',             12, '类型'),
        ('F', '所属模块',         22, '所属模块'),
        ('G', '前置条件',         40, '前置条件'),
        ('H', '执行步骤（含预期结果）', 70, '执行步骤'),
        ('I', '后置条件',         30, '后置条件'),
    ]

    # ── 表头 ─────────────────────────────────────────────────────────
    for idx, (col, header, width, _) in enumerate(columns, start=1):
        ws.column_dimensions[col].width = width
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = thin_border

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    # ── 数据行 ───────────────────────────────────────────────────────
    for row_idx, tc in enumerate(test_cases, start=2):
        priority = tc.get('优先级', '')

        for col_idx, (col, _, _, key) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=tc.get(key, ''))
            cell.border = thin_border

            # 优先级列：居中 + 着色
            if key == '优先级':
                cell.font      = Font(name='微软雅黑', size=10, bold=True)
                cell.alignment = center_align
                if priority == 'P0':
                    cell.fill = fill_p0
                elif priority == 'P1':
                    cell.fill = fill_p1
                elif priority == 'P2':
                    cell.fill = fill_p2

            # 粒度 / 类型：居中
            elif key in ('粒度', '类型'):
                cell.font      = default_font
                cell.alignment = center_align

            # 所属模块：蓝色加粗
            elif key == '所属模块':
                cell.font      = module_font
                cell.alignment = default_align

            # 其余列：默认样式
            else:
                cell.font      = default_font
                cell.alignment = default_align

    # 保存
    wb.save(output_file)
    print(f'✓ Excel文件生成成功: {output_file}')
    print(f'  共 {len(test_cases)} 条测试用例')

    # 优先级统计
    from collections import Counter
    prio_count = Counter(tc.get('优先级', '') for tc in test_cases)
    for p in ('P0', 'P1', 'P2', 'P3'):
        if prio_count.get(p):
            print(f'  {p}: {prio_count[p]} 条')
    return True


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('使用方法: python md_to_excel.py <markdown文件> [输出文件]')
        print('示例: python md_to_excel.py docs/测试用例_数据模型V0.1.md')
        sys.exit(1)

    md_file    = sys.argv[1]
    excel_file = sys.argv[2] if len(sys.argv) > 2 else md_file.replace('.md', '.xlsx')

    test_cases = parse_test_case_markdown(md_file)

    if not test_cases:
        print(f'⚠️  未解析到任何测试用例，请确认文件格式是否正确: {md_file}')
        sys.exit(1)

    create_test_case_excel(test_cases, excel_file)
